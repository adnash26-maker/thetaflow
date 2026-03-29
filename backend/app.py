"""
ThetaFlow - Investment Intelligence Platform

Ingests current events, maps them to value chains, simulates impact
propagation, and generates investment recommendations across the
full supply chain — from raw materials to end applications.
"""

import os
import sys
import io
import csv
import json
import logging
import sqlite3
import re
import time
import uuid
import requests
import hashlib
import secrets
from datetime import datetime, timedelta

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))

from flask import Flask, jsonify, request, send_from_directory, Response, redirect, session
from flask_cors import CORS
from functools import wraps

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from event_ingestion import EventDatabase, EventOrchestrator
from impact_engine import ImpactEngine
from value_chains import ALL_CHAINS, find_chains_for_event, get_chain_tickers
from stock_universe import StockUniverse, load_sec_with_sic
from ai_analyst import AIAnalyst

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("thetaflow.api")

_frontend_dir = os.getenv("THETAFLOW_FRONTEND", os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "frontend"))
app = Flask(__name__, static_folder=_frontend_dir, static_url_path="")
app.secret_key = os.getenv("FLASK_SECRET_KEY", secrets.token_hex(32))
CORS(app, supports_credentials=True)

DB_PATH = os.getenv("THETAFLOW_DB", os.path.join(os.path.expanduser("~"), "thetaflow.db"))

db = EventDatabase(DB_PATH)
orchestrator = EventOrchestrator(db)
engine = ImpactEngine(DB_PATH)
universe = StockUniverse(DB_PATH)
analyst = AIAnalyst()

last_collection_at = None

# ── Auth (reused from TrendSniper) ──

def hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac('sha256', password.encode(), salt.encode(), 100000)
    return f"{salt}:{h.hex()}"

def verify_password(password: str, stored: str) -> bool:
    salt, h = stored.split(':')
    return hashlib.pbkdf2_hmac('sha256', password.encode(), salt.encode(), 100000).hex() == h

def auth_required(tier_minimum=None):
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            user = None
            api_key = request.headers.get('X-API-Key')
            if api_key:
                conn = sqlite3.connect(DB_PATH)
                conn.row_factory = sqlite3.Row
                row = conn.execute("SELECT * FROM users WHERE api_key = ?", (api_key,)).fetchone()
                conn.close()
                if row:
                    user = dict(row)
            if not user and 'user_id' in session:
                conn = sqlite3.connect(DB_PATH)
                conn.row_factory = sqlite3.Row
                row = conn.execute("SELECT * FROM users WHERE id = ?", (session['user_id'],)).fetchone()
                conn.close()
                if row:
                    user = dict(row)
            if not user:
                return jsonify({"error": "Authentication required"}), 401
            if tier_minimum:
                tiers = {'free': 0, 'starter': 1, 'pro': 2, 'enterprise': 3}
                if tiers.get(user.get('tier', 'free'), 0) < tiers.get(tier_minimum, 0):
                    return jsonify({"error": f"Requires {tier_minimum} tier"}), 403
            request.current_user = user
            return f(*args, **kwargs)
        return wrapped
    return decorator

# ── Pages ──

def _find_html(filename):
    """Find and read an HTML file - tries every possible path.
    Priority: THETAFLOW_FRONTEND env > backend dir > frontend sibling > cwd."""
    _this_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.getenv("THETAFLOW_FRONTEND", ""),
        _this_dir,  # backend/ dir (has copies)
        os.path.join(_this_dir, "..", "frontend"),  # ../frontend/
        os.getcwd(),
        os.path.join(os.getcwd(), "frontend"),
        os.path.join(os.getcwd(), "backend"),
        "/app/frontend",  # Railway common path
        "/app/backend",
    ]
    for d in candidates:
        if not d:
            continue
        path = os.path.join(d, filename)
        if os.path.isfile(path):
            logger.debug(f"Found {filename} at {path}")
            with open(path, "r") as f:
                return f.read()
    logger.error(f"Could not find {filename} in any candidate path: {candidates}")
    return None

@app.route("/")
def serve_landing():
    html = _find_html("landing.html")
    if html:
        return Response(html, mimetype="text/html")
    return "Landing page not found", 404

@app.route("/dashboard")
def serve_dashboard():
    html = _find_html("dashboard.html")
    if html:
        return Response(html, mimetype="text/html")
    return "Dashboard not found", 404

# ── Core API ──

@app.route("/api/debug-paths")
def debug_paths():
    """Debug route to find frontend directory on Railway."""
    import glob
    _this_dir = os.path.dirname(os.path.abspath(__file__))
    cwd = os.getcwd()
    info = {
        "cwd": cwd,
        "app_file": os.path.abspath(__file__),
        "this_dir": _this_dir,
        "env_frontend": os.getenv("THETAFLOW_FRONTEND", "not set"),
        "cwd_contents": sorted(os.listdir(cwd)) if os.path.isdir(cwd) else [],
        "this_dir_contents": sorted(os.listdir(_this_dir)) if os.path.isdir(_this_dir) else [],
        "dashboard_found": _find_html("dashboard.html") is not None,
        "landing_found": _find_html("landing.html") is not None,
    }
    # Search common paths
    for search_dir in [cwd, _this_dir, "/app", os.path.join(cwd, "..")]:
        try:
            for f in glob.glob(os.path.join(search_dir, "**", "dashboard.html"), recursive=True):
                info[f"found_{f}"] = True
        except Exception:
            pass
    return jsonify(info)

@app.route("/api/health")
def health():
    return jsonify({
        "status": "ok",
        "product": "ThetaFlow",
        "timestamp": datetime.utcnow().isoformat(),
        "last_collection": last_collection_at,
        "chains_available": len(ALL_CHAINS),
    })

@app.route("/api/chains", methods=["GET"])
def get_chains():
    """List all available value chains."""
    chains = []
    for cid, chain in ALL_CHAINS.items():
        total_tickers = sum(len(n.tickers) for n in chain.nodes)
        chains.append({
            "id": cid,
            "name": chain.name,
            "description": chain.description,
            "theme_color": chain.theme_color,
            "node_count": len(chain.nodes),
            "ticker_count": total_tickers,
            "catalyst_keywords": chain.catalyst_keywords[:5],
        })
    return jsonify({"success": True, "chains": chains})

@app.route("/api/chains/<chain_id>", methods=["GET"])
def get_chain_detail(chain_id):
    """Get full analysis for a specific value chain with live prices."""
    analysis = engine.get_chain_analysis(chain_id)
    if "error" in analysis:
        return jsonify({"success": False, "error": analysis["error"]}), 404
    return jsonify({"success": True, **analysis})

@app.route("/api/top-headlines", methods=["GET"])
def get_top_headlines():
    """Get live headlines from multiple RSS sources, one per category."""
    import xml.etree.ElementTree as ET
    from concurrent.futures import ThreadPoolExecutor

    # Multiple feeds per category for redundancy. First working feed wins.
    feed_groups = {
        "Technology": {
            "color": "#22d3ee",
            "feeds": [
                "https://news.google.com/rss/topics/CAAqJggKIiBDQkFTRWdvSUwyMHZNRGx6TVdZU0FtVnVHZ0pWVXlnQVAB?hl=en-US&gl=US&ceid=US:en",
            ],
        },
        "Markets": {
            "color": "#34d399",
            "feeds": [
                "https://finance.yahoo.com/news/rssindex",
                "https://feeds.marketwatch.com/marketwatch/topstories/",
            ],
        },
        "Business": {
            "color": "#818cf8",
            "feeds": [
                "https://news.google.com/rss/topics/CAAqJggKIiBDQkFTRWdvSUwyMHZNRGx6TVdZU0FtVnVHZ0pWVXlnQVAB?hl=en-US&gl=US&ceid=US:en",
                "https://feeds.marketwatch.com/marketwatch/topstories/",
            ],
        },
        "Economy": {
            "color": "#fbbf24",
            "feeds": [
                "https://news.google.com/rss/topics/CAAqIggKIhxDQkFTRHdvSkwyMHZNR2RtZUhBU0FtVnVLQUFQAQ?hl=en-US&gl=US&ceid=US:en",
                "https://finance.yahoo.com/news/rssindex",
            ],
        },
    }

    def fetch_feed(url):
        """Fetch and parse a single RSS feed, return list of items."""
        try:
            resp = requests.get(url, headers={
                "User-Agent": "Mozilla/5.0 (compatible; ThetaFlow/2.0)",
                "Cache-Control": "no-cache",
            }, timeout=8)
            if resp.status_code != 200:
                return []
            root = ET.fromstring(resp.content)
            items = root.findall(".//item")
            results = []
            for item in items[:10]:
                title = (item.findtext("title") or "").strip()
                link = (item.findtext("link") or "").strip()
                if title and len(title) > 15:
                    results.append({"title": title[:120], "url": link})
            return results
        except Exception:
            return []

    # Fetch all unique feed URLs in parallel
    all_urls = set()
    for group in feed_groups.values():
        all_urls.update(group["feeds"])

    feed_cache = {}
    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = {url: executor.submit(fetch_feed, url) for url in all_urls}
        for url, future in futures.items():
            try:
                feed_cache[url] = future.result(timeout=10)
            except Exception:
                feed_cache[url] = []

    # Pick one headline per category, avoid duplicates
    headlines = []
    seen_titles = set()
    # Track which item index to use per feed to avoid reusing the same headline
    feed_idx = {url: 0 for url in all_urls}

    for category, info in feed_groups.items():
        found = False
        for feed_url in info["feeds"]:
            items = feed_cache.get(feed_url, [])
            idx = feed_idx.get(feed_url, 0)
            while idx < len(items):
                item = items[idx]
                idx += 1
                title_key = item["title"][:50].lower()
                if title_key not in seen_titles:
                    seen_titles.add(title_key)
                    headlines.append({
                        "category": category,
                        "title": item["title"],
                        "url": item["url"],
                        "color": info["color"],
                        "source": "Live",
                    })
                    feed_idx[feed_url] = idx
                    found = True
                    break
            if found:
                break

    response = jsonify({"success": True, "headlines": headlines})
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

@app.route("/api/signals", methods=["GET"])
def get_signals():
    """Get today's top events pre-analyzed with chain impacts and ticker recommendations.
    This is the primary dashboard endpoint — shows what happened and what it means."""
    hours = request.args.get("hours", 72, type=int)
    limit = request.args.get("limit", 10, type=int)

    events = db.get_recent_events(hours=hours, limit=50)
    # Prioritize WSJ and news headlines over SEC filings
    priority = {"wsj": 0, "newsapi": 1, "fred": 2, "sec_edgar": 3, "yahoo": 4}
    events.sort(key=lambda e: (priority.get(e.get("source", ""), 5), e.get("timestamp", "")), reverse=False)
    events.sort(key=lambda e: priority.get(e.get("source", ""), 5))
    signals = []

    for event in events:
        try:
            matched_chains = json.loads(event.get("matched_chains", "[]"))
        except (json.JSONDecodeError, TypeError):
            matched_chains = []

        # Only show events that matched at least one chain
        if not matched_chains:
            # Try re-analyzing (event might have been stored before keyword improvements)
            matched_chains = find_chains_for_event(event["title"])

        if not matched_chains:
            continue

        # Get top ticker recommendations for this event
        analysis = engine.analyze_event(event["title"], event.get("event_type", "news"))
        top_picks = analysis.get("top_picks", [])[:3]

        signals.append({
            "title": event["title"],
            "source": event["source"],
            "severity": event["severity"],
            "timestamp": event["timestamp"],
            "url": event.get("url", ""),
            "chains": matched_chains[:2],
            "top_picks": top_picks,
            "summary": analysis.get("summary", ""),
        })

        if len(signals) >= limit:
            break

    return jsonify({
        "success": True,
        "signals": signals,
        "count": len(signals),
        "generated_at": datetime.utcnow().isoformat(),
    })

@app.route("/api/analyze-stream", methods=["POST"])
def analyze_stream():
    """Stream analysis results via SSE for progressive rendering."""
    from financial_data import get_full_financials_batch, get_ticker_chart_data
    from concurrent.futures import ThreadPoolExecutor

    data = request.json or {}
    headline = data.get("headline", "").strip()
    if not headline:
        return jsonify({"error": "headline is required"}), 400

    def sse(event_type, payload):
        return f"event: {event_type}\ndata: {json.dumps(payload)}\n\n"

    def generate():
        yield sse("status", {"message": "Identifying investment themes...", "phase": "ai", "pct": 5})

        if not analyst.available:
            yield sse("error", {"message": "AI analyst not available"})
            return

        try:
            # Stream Claude API — send progress as tokens arrive
            prompt = analyst._build_dynamic_prompt(headline)
            full_text = ""

            with analyst.client.messages.stream(
                model="claude-sonnet-4-20250514",
                max_tokens=6000,
                messages=[{"role": "user", "content": prompt}]
            ) as stream:
                token_count = 0
                last_pct = 5
                for chunk in stream.text_stream:
                    full_text += chunk
                    token_count += 1
                    # Send progress every ~40 tokens (~10 updates total)
                    pct = min(5 + int(token_count / 5), 85)
                    if pct >= last_pct + 8:
                        last_pct = pct
                        messages = [
                            "Analyzing supply chain connections...",
                            "Finding historical precedents...",
                            "Computing earnings impact...",
                            "Identifying cross-catalyst compounding...",
                            "Scoring conviction levels...",
                            "Ranking non-obvious plays...",
                            "Building trade recommendations...",
                            "Quantifying risk/reward...",
                            "Finalizing analysis...",
                        ]
                        msg_idx = min(pct // 10, len(messages) - 1)
                        yield sse("progress", {"pct": pct, "message": messages[msg_idx]})

            # Parse the complete response
            ai_result = analyst._parse_dynamic_result(full_text)
            if not ai_result or not ai_result.get("top_picks"):
                logger.error(f"Stream parse failed. Text length: {len(full_text)}, starts with: {full_text[:200]}")
                # Fallback: try the non-streaming path
                try:
                    ai_result = analyst.generate_dynamic_chains(headline)
                except Exception:
                    ai_result = None
                if not ai_result or not ai_result.get("top_picks"):
                    yield sse("error", {"message": "Could not generate analysis. Try a more specific headline."})
                    return

            # Send raw AI result immediately (before financial enrichment)
            yield sse("ai_result", {
                "chains": ai_result.get("chains", []),
                "summary": ai_result.get("summary", ""),
                "obvious_play": ai_result.get("obvious_play"),
                "historical_precedents": ai_result.get("historical_precedents", []),
                "active_catalysts": ai_result.get("active_catalysts", []),
                "top_picks_preview": [
                    {"ticker": p.get("ticker"), "company": p.get("company"),
                     "direction": p.get("direction"), "conviction": p.get("conviction"),
                     "action": p.get("action"), "order": p.get("order"),
                     "thesis": p.get("thesis"), "earnings_impact": p.get("earnings_impact"),
                     "precedent_move": p.get("precedent_move"), "layer": p.get("layer"),
                     "chain_name": p.get("chain_name"), "exposure": p.get("exposure"),
                     "time_horizon": p.get("time_horizon"), "impact_score": p.get("impact_score"),
                     "risk_reward": p.get("risk_reward")}
                    for p in ai_result.get("top_picks", [])
                ],
                "risk_factors": ai_result.get("risk_factors", []),
                "contrarian_view": ai_result.get("contrarian_view", ""),
                "time_horizon": ai_result.get("time_horizon", ""),
            })

            yield sse("status", {"message": "Fetching live market data...", "phase": "enrich", "pct": 88})

            # Enrich with financial data
            enriched = _enrich_dynamic_result(ai_result, headline)
            enriched["ai_powered"] = True
            enriched["dynamic_chains"] = True

            _save_recommendations(enriched)

            yield sse("complete", {"success": True, **enriched})

        except Exception as e:
            logger.error(f"Stream analysis failed: {e}")
            yield sse("error", {"message": f"Analysis failed: {str(e)[:300]}"})

    return Response(
        generate(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        }
    )


@app.route("/api/analyze", methods=["POST"])
def analyze_event_endpoint():
    """Analyze a news headline with AI-powered dynamic value chain generation."""
    from financial_data import get_full_financials, get_ticker_chart_data

    data = request.json or {}
    headline = data.get("headline", "").strip()
    if not headline:
        return jsonify({"error": "headline is required"}), 400

    event_type = data.get("event_type", "news")
    analysis = None

    # ── Primary path: Dynamic AI-generated chains ──
    if analyst.available:
        try:
            ai_result = analyst.generate_dynamic_chains(headline)
            if ai_result and ai_result.get("top_picks"):
                analysis = _enrich_dynamic_result(ai_result, headline)
                analysis["ai_powered"] = True
                analysis["dynamic_chains"] = True
        except Exception as e:
            logger.error(f"Dynamic chain generation failed, falling back: {e}")

    # ── Fallback path: Pre-built chain matching ──
    if analysis is None:
        analysis = engine.analyze_event(headline, event_type)

        if analyst.available and analysis.get("chains_matched", 0) > 0:
            try:
                ticker_data = []
                for t in analysis.get("top_picks", [])[:5]:
                    ticker_data.append({
                        "ticker": t["ticker"], "company": t["company"],
                        "current_price": t.get("current_price", 0),
                        "change_pct": t.get("change_pct", 0),
                        "chain_name": t.get("chain_name", ""),
                        "layer": t.get("layer", ""),
                        "exposure": t.get("exposure", ""),
                        "thesis": t.get("thesis", ""),
                        "financials": t.get("financials", {}),
                    })
                ai_text = analyst.analyze_event(headline, analysis.get("chains", []), ticker_data)
                if ai_text.get("summary"):
                    analysis["summary"] = ai_text["summary"]
                if ai_text.get("risk_factors"):
                    analysis["risk_factors"] = ai_text["risk_factors"]
                if ai_text.get("contrarian_view"):
                    analysis["contrarian_view"] = ai_text["contrarian_view"]
                if ai_text.get("time_horizon"):
                    analysis["ai_time_horizon"] = ai_text["time_horizon"]
                ticker_analyses = ai_text.get("ticker_analyses", {})
                for pick in analysis.get("top_picks", []):
                    ai_t = ticker_analyses.get(pick["ticker"])
                    if ai_t:
                        pick["investment_analysis"] = ai_t
                analysis["ai_powered"] = True
            except Exception:
                analysis["ai_powered"] = False
        else:
            analysis["ai_powered"] = False

        # Add chart data to fallback picks
        _add_chart_data_to_picks(analysis.get("top_picks", []))
        analysis["dynamic_chains"] = False

    # Save top picks to recommendation history for scorecard tracking
    _save_recommendations(analysis)

    return jsonify({"success": True, **analysis})


def _save_recommendations(analysis: dict):
    """Save top picks to DB for performance tracking."""
    try:
        headline = analysis.get("headline", "")
        conn = sqlite3.connect(DB_PATH)
        for pick in analysis.get("top_picks", [])[:5]:
            if not pick.get("ticker") or not pick.get("current_price"):
                continue
            # Compute target price
            proj_sign = 1 if pick.get("direction") == "bullish" else -1
            conv = pick.get("conviction", 0.5)
            import math
            target = pick["current_price"] * math.exp(proj_sign * conv * 0.003 * 22)
            conn.execute("""
                INSERT INTO recommendation_history
                (ticker, company, action, direction, conviction, entry_price, target_price,
                 headline, order_type, chain_name, risk_reward)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                pick["ticker"], pick.get("company", ""),
                pick.get("action", "BUY"), pick.get("direction", "bullish"),
                conv, pick["current_price"], round(target, 2),
                headline[:200], pick.get("order", 1),
                pick.get("chain_name", ""), pick.get("risk_reward", ""),
            ))
        conn.commit()
        conn.close()
    except Exception as e:
        logger.debug(f"Failed to save recommendations: {e}")


@app.route("/api/scorecard", methods=["GET"])
def get_scorecard():
    """Get performance scorecard of past recommendations with live P&L."""
    from financial_data import _get_basic_quote

    days = request.args.get("days", 14, type=int)
    limit = request.args.get("limit", 10, type=int)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cutoff = (datetime.utcnow() - timedelta(days=days)).isoformat()

    # Get unique recent recommendations (dedupe by ticker, keep latest)
    rows = conn.execute("""
        SELECT * FROM recommendation_history
        WHERE created_at > ? ORDER BY created_at DESC LIMIT ?
    """, (cutoff, limit * 3)).fetchall()
    conn.close()

    seen = set()
    recs = []
    for row in rows:
        r = dict(row)
        if r["ticker"] in seen:
            continue
        seen.add(r["ticker"])

        # Fetch current price
        quote = _get_basic_quote(r["ticker"])
        if quote and quote.get("price"):
            current = quote["price"]
            entry = r["entry_price"]
            if r["direction"] == "bearish":
                pnl_pct = round((entry - current) / entry * 100, 1)
            else:
                pnl_pct = round((current - entry) / entry * 100, 1)

            recs.append({
                "ticker": r["ticker"],
                "company": r["company"],
                "action": r["action"],
                "direction": r["direction"],
                "conviction": r["conviction"],
                "entry_price": entry,
                "current_price": current,
                "target_price": r["target_price"],
                "pnl_pct": pnl_pct,
                "correct": pnl_pct > 0,
                "headline": r["headline"],
                "order_type": r["order_type"],
                "chain_name": r["chain_name"],
                "created_at": r["created_at"],
            })
            time.sleep(0.2)

        if len(recs) >= limit:
            break

    wins = sum(1 for r in recs if r["correct"])
    total = len(recs)

    return jsonify({
        "success": True,
        "recommendations": recs,
        "stats": {
            "total": total,
            "wins": wins,
            "win_rate": round(wins / total * 100) if total > 0 else 0,
            "avg_pnl": round(sum(r["pnl_pct"] for r in recs) / total, 1) if total > 0 else 0,
        }
    })


@app.route("/api/scored-picks", methods=["GET"])
def get_scored_picks():
    """Get AI-scored stock picks across all chains with ThetaFlow Score."""
    from financial_data import get_full_financials

    # 1. Gather all unique tickers from chains + recommendation history
    all_tickers = {}
    for chain in ALL_CHAINS.values():
        for node in chain.nodes:
            for t in node.tickers:
                tk = t["ticker"]
                if tk not in all_tickers:
                    all_tickers[tk] = {
                        "ticker": tk, "company": t["company"],
                        "chain_name": chain.name, "layer": node.layer,
                        "exposure": t.get("exposure", "medium"),
                        "sensitivity": node.sensitivity,
                    }

    # 2. Get recommendation frequency from history
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("""
        SELECT ticker, COUNT(*) as freq, AVG(conviction) as avg_conv,
               MAX(created_at) as last_seen
        FROM recommendation_history GROUP BY ticker
    """).fetchall()
    conn.close()
    rec_freq = {r["ticker"]: dict(r) for r in rows}

    # 3. Enrich top candidates with live prices
    candidates = sorted(all_tickers.values(),
                       key=lambda t: rec_freq.get(t["ticker"], {}).get("avg_conv", t["sensitivity"]),
                       reverse=True)[:30]

    enriched = []
    for c in candidates:
        fin = get_full_financials(c["ticker"])
        if not fin or not fin.get("price"):
            continue
        c["current_price"] = fin["price"]
        c["change_pct"] = fin.get("change_pct", 0)
        c["market_cap"] = fin.get("market_cap_str", "")
        c["pe_forward"] = fin.get("pe_forward")
        c["distance_from_high"] = fin.get("distance_from_high_pct")
        c["beta"] = fin.get("beta")
        rf = rec_freq.get(c["ticker"], {})
        c["rec_frequency"] = rf.get("freq", 0)
        c["avg_conviction"] = rf.get("avg_conv", c["sensitivity"])
        enriched.append(c)
        time.sleep(0.2)
        if len(enriched) >= 20:
            break

    # 4. Send to Claude for AI scoring
    if analyst.available and enriched:
        try:
            ticker_list = "\n".join([
                f"- {t['ticker']} ({t['company']}): ${t['current_price']:.0f}, "
                f"chain={t['chain_name']}, P/E={t.get('pe_forward') or 'N/A'}, "
                f"from_high={t.get('distance_from_high') or 'N/A'}%, "
                f"rec_freq={t['rec_frequency']}, avg_conv={t['avg_conviction']:.0%}"
                for t in enriched
            ])

            response = analyst.client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=3000,
                messages=[{"role": "user", "content": f"""Score these stocks on a 1-100 ThetaFlow Score based on: catalyst exposure, valuation attractiveness, momentum, and risk/reward asymmetry. Consider current market conditions.

STOCKS:
{ticker_list}

Return JSON array. For each stock:
{{"ticker":"SYM","score":85,"grade":"A","action":"BUY","rationale":"One sentence why."}}

Score meaning: 90-100=Strong Buy, 75-89=Buy, 60-74=Watch, 40-59=Neutral, below 40=Avoid.
Grade: A/B/C/D/F.
Sort by score descending. Return ONLY the JSON array, no other text."""}]
            )

            text = response.content[0].text.strip()
            if "```json" in text:
                text = text.split("```json")[1].split("```")[0].strip()
            elif "```" in text:
                text = text.split("```")[1].split("```")[0].strip()

            scores = json.loads(text)
            score_map = {s["ticker"]: s for s in scores if isinstance(s, dict)}

            for t in enriched:
                sc = score_map.get(t["ticker"], {})
                t["score"] = sc.get("score", 50)
                t["grade"] = sc.get("grade", "C")
                t["action"] = sc.get("action", "WATCH")
                t["rationale"] = sc.get("rationale", "")

        except Exception as e:
            logger.error(f"AI scoring failed: {e}")
            for t in enriched:
                t["score"] = round(t["avg_conviction"] * 100)
                t["grade"] = "A" if t["score"] >= 75 else "B" if t["score"] >= 60 else "C"
                t["action"] = "BUY" if t["score"] >= 75 else "WATCH"
                t["rationale"] = ""
    else:
        for t in enriched:
            t["score"] = round(t["avg_conviction"] * 100)
            t["grade"] = "A" if t["score"] >= 75 else "B" if t["score"] >= 60 else "C"
            t["action"] = "BUY" if t["score"] >= 75 else "WATCH"
            t["rationale"] = ""

    enriched.sort(key=lambda t: t.get("score", 0), reverse=True)

    return jsonify({
        "success": True,
        "picks": enriched,
        "total": len(enriched),
        "ai_scored": analyst.available,
    })


def _enrich_dynamic_result(ai_result: dict, headline: str) -> dict:
    """Enrich Claude's dynamic chain result with live financial data and charts.
    Uses parallel batch fetching for speed."""
    from financial_data import get_full_financials_batch, get_ticker_chart_data
    from concurrent.futures import ThreadPoolExecutor

    all_picks = ai_result.get("top_picks", []) + [
        r for r in ai_result.get("recommendations", [])
        if r.get("ticker") not in {p.get("ticker") for p in ai_result.get("top_picks", [])}
    ]

    # Collect unique valid tickers
    unique_tickers = []
    seen = set()
    for pick in all_picks:
        ticker = pick.get("ticker", "").upper().strip()
        if ticker and ticker not in seen and len(ticker) <= 5:
            seen.add(ticker)
            unique_tickers.append(ticker)

    # Parallel batch fetch all financials at once
    batch_data = get_full_financials_batch(unique_tickers[:12])

    # Pre-fetch chart data in parallel for top 5 tickers
    chart_tickers = unique_tickers[:5]
    chart_data_map = {}
    if chart_tickers:
        def _fetch_chart(tk):
            # Find the pick to get direction/conviction
            for p in all_picks:
                if p.get("ticker", "").upper().strip() == tk:
                    return (tk, get_ticker_chart_data(tk, p.get("direction", "bullish"), p.get("conviction", 0.5)))
            return (tk, get_ticker_chart_data(tk))

        with ThreadPoolExecutor(max_workers=5) as executor:
            for tk, chart in executor.map(lambda t: _fetch_chart(t), chart_tickers):
                if chart:
                    chart_data_map[tk] = chart

    valid_picks = []
    seen_tickers = set()

    for pick in all_picks:
        ticker = pick.get("ticker", "").upper().strip()
        if not ticker or ticker in seen_tickers or len(ticker) > 5:
            continue

        fin = batch_data.get(ticker)
        if fin is None:
            logger.warning(f"Ticker {ticker} not found on Yahoo Finance, skipping")
            continue

        seen_tickers.add(ticker)

        pick["current_price"] = fin.get("price", 0)
        pick["change_pct"] = fin.get("change_pct", 0)
        pick["financials"] = {
            "market_cap": fin.get("market_cap_str"),
            "pe_forward": fin.get("pe_forward"),
            "pe_trailing": fin.get("pe_trailing"),
            "revenue_growth": fin.get("revenue_growth"),
            "profit_margin": fin.get("profit_margin"),
            "fifty_two_high": fin.get("fifty_two_high"),
            "fifty_two_low": fin.get("fifty_two_low"),
            "distance_from_high_pct": fin.get("distance_from_high_pct"),
            "beta": fin.get("beta"),
        }

        pick["investment_analysis"] = pick.get("thesis", "")

        # Attach pre-fetched chart data
        if ticker in chart_data_map:
            pick["chart_data"] = chart_data_map[ticker]

        valid_picks.append(pick)

    all_valid = sorted(
        valid_picks,
        key=lambda t: t.get("conviction", 0) * t.get("impact_score", 0),
        reverse=True
    )

    return {
        "headline": headline,
        "event_type": "news",
        "chains_matched": len(ai_result.get("chains", [])),
        "chains": ai_result.get("chains", []),
        "top_picks": all_valid[:5],
        "recommendations": all_valid,
        "summary": ai_result.get("summary", ""),
        "risk_factors": ai_result.get("risk_factors", []),
        "obvious_play": ai_result.get("obvious_play"),
        "historical_precedents": ai_result.get("historical_precedents", []),
        "active_catalysts": ai_result.get("active_catalysts", []),
        "contrarian_view": ai_result.get("contrarian_view", ""),
        "ai_time_horizon": ai_result.get("time_horizon", ""),
        "analyzed_at": datetime.utcnow().isoformat(),
    }


def _add_chart_data_to_picks(picks):
    """Add chart data to ticker picks (fallback path)."""
    from financial_data import get_ticker_chart_data

    for pick in (picks or [])[:5]:
        ticker = pick.get("ticker")
        if not ticker:
            continue
        chart = get_ticker_chart_data(
            ticker,
            direction=pick.get("direction", "bullish"),
            conviction=pick.get("conviction", 0.5)
        )
        if chart:
            pick["chart_data"] = chart

@app.route("/api/events", methods=["GET"])
def get_events():
    """Get recent catalyst events with their chain impacts."""
    hours = request.args.get("hours", 72, type=int)
    limit = request.args.get("limit", 30, type=int)
    events = db.get_recent_events(hours=hours, limit=limit)
    # Parse JSON fields
    for e in events:
        try:
            e["matched_chains"] = json.loads(e.get("matched_chains", "[]"))
            e["metadata"] = json.loads(e.get("metadata", "{}"))
        except (json.JSONDecodeError, TypeError):
            e["matched_chains"] = []
            e["metadata"] = {}
    return jsonify({"success": True, "events": events, "count": len(events)})

@app.route("/api/portfolio", methods=["GET"])
def get_portfolio():
    """Get cross-chain portfolio view of all recommended tickers."""
    portfolio = engine.get_portfolio_view()
    return jsonify({"success": True, **portfolio})

@app.route("/api/ticker/<ticker>", methods=["GET"])
def get_ticker_info(ticker):
    """Get info about a specific ticker including which chains it belongs to."""
    ticker = ticker.upper()
    chains_containing = []
    for cid, chain in ALL_CHAINS.items():
        for node in chain.nodes:
            for t in node.tickers:
                if t["ticker"] == ticker:
                    chains_containing.append({
                        "chain_id": cid,
                        "chain_name": chain.name,
                        "layer": node.layer,
                        "position": node.position,
                        "sensitivity": node.sensitivity,
                        "exposure": t.get("exposure", "medium"),
                        "notes": t.get("notes", ""),
                    })

    if not chains_containing:
        return jsonify({"success": False, "error": "Ticker not found in any chain"}), 404

    # Get price history
    history = db.get_ticker_history(ticker, days=30)

    return jsonify({
        "success": True,
        "ticker": ticker,
        "company": chains_containing[0].get("notes", "").split(".")[0] if chains_containing else "",
        "chains": chains_containing,
        "chain_count": len(chains_containing),
        "price_history": history,
    })

@app.route("/api/universe/search", methods=["GET"])
def search_universe():
    """Search all publicly listed stocks."""
    query = request.args.get("q", "")
    chain_id = request.args.get("chain_id")
    limit = request.args.get("limit", 20, type=int)
    results = universe.search_tickers(query, chain_id, limit)
    return jsonify({"success": True, "results": results, "count": len(results)})

@app.route("/api/universe/stats", methods=["GET"])
def universe_stats():
    """Get stock universe coverage stats."""
    stats = universe.get_universe_stats()
    return jsonify({"success": True, **stats})

@app.route("/api/universe/chain/<chain_id>", methods=["GET"])
def universe_chain_companies(chain_id):
    """Get all companies in the universe mapped to a chain."""
    layer = request.args.get("layer")
    limit = request.args.get("limit", 50, type=int)
    companies = universe.get_chain_companies(chain_id, layer, limit)
    return jsonify({"success": True, "companies": companies, "count": len(companies)})

@app.route("/api/universe/load", methods=["POST"])
def load_universe():
    """Load/refresh the stock universe from SEC EDGAR."""
    count = universe.load_from_sec()
    return jsonify({"success": True, "companies_loaded": count})

@app.route("/api/collect", methods=["POST"])
def trigger_collection():
    """Manually trigger event collection."""
    global last_collection_at
    try:
        counts = orchestrator.run_collection()
        last_collection_at = datetime.utcnow().isoformat()
        return jsonify({"success": True, "collected": counts})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/export", methods=["GET"])
def export_csv_portfolio():
    """Export portfolio recommendations as CSV."""
    portfolio = engine.get_portfolio_view()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ticker", "company", "chain", "layer", "direction",
                     "conviction", "impact_score", "time_horizon", "price", "change_pct", "thesis"])
    for t in portfolio.get("tickers", []):
        writer.writerow([
            t["ticker"], t["company"], t["chain_name"], t["layer"],
            t["direction"], t["conviction"], t["impact_score"],
            t["time_horizon"], t["current_price"], t["change_pct"], t["thesis"]
        ])
    return Response(output.getvalue(), mimetype='text/csv',
                    headers={"Content-Disposition": "attachment;filename=thetaflow_portfolio.csv"})

@app.route("/api/export-xlsx", methods=["POST"])
def export_xlsx():
    """Export analysis as Excel workbook with DCF, revenue buildup, P&L pro forma."""
    import math
    from openpyxl.utils import get_column_letter as CL
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500

    data = request.json or {}
    if not data.get("top_picks"):
        return jsonify({"error": "No analysis data provided"}), 400

    wb = Workbook()

    # ── Styles ──
    hdr_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    hdr_fill = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
    input_fill = PatternFill(start_color='FFF9E6', end_color='FFF9E6', fill_type='solid')  # yellow = editable
    sec_font = Font(name='Calibri', bold=True, size=11, color='1B2A4A')
    sec_font_sm = Font(name='Calibri', bold=True, size=10, color='1B2A4A')
    label_font = Font(name='Calibri', size=10, color='4A5568')
    num_font = Font(name='Consolas', size=10)
    num_font_b = Font(name='Consolas', bold=True, size=10)
    link_font = Font(name='Consolas', size=10, color='2563EB')
    green_font = Font(name='Consolas', bold=True, size=10, color='10B981')
    red_font = Font(name='Consolas', bold=True, size=10, color='EF4444')
    pct_fmt = '0.0%'
    pct1_fmt = '0.0%'
    usd_fmt = '$#,##0.00'
    usd_m = '$#,##0'
    usd_whole = '$#,##0'
    bps_fmt = '#,##0'
    thin_border = Border(bottom=Side(style='thin', color='D0D5DD'))
    thick_border = Border(
        top=Side(style='medium', color='1B2A4A'),
        bottom=Side(style='medium', color='1B2A4A'),
    )

    def hdr_row(ws, row, col_count):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def auto_w(ws):
        for col in ws.columns:
            mx = 0
            cl = col[0].column_letter
            for cell in col:
                if cell.value:
                    mx = max(mx, len(str(cell.value)))
            ws.column_dimensions[cl].width = min(mx + 3, 50)

    def write_input(ws, row, col, value, fmt=None):
        """Write a yellow (editable) input cell."""
        c = ws.cell(row=row, column=col, value=value)
        c.fill = input_fill
        c.font = num_font
        if fmt:
            c.number_format = fmt
        return c

    def write_formula(ws, row, col, formula, fmt=None, bold=False):
        """Write a formula cell."""
        c = ws.cell(row=row, column=col, value=formula)
        c.font = num_font_b if bold else num_font
        if fmt:
            c.number_format = fmt
        return c

    def write_label(ws, row, col, text, bold=False, indent=0):
        c = ws.cell(row=row, column=col, value=("  " * indent) + text)
        c.font = sec_font_sm if bold else label_font
        return c

    top_picks = data.get("top_picks", [])
    all_picks = data.get("recommendations", top_picks)
    PROJ_YEARS = 5

    # ════════════════════════════════════════
    # SHEET 1: Executive Summary
    # ════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "22D3EE"
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 50

    ws.cell(row=1, column=1, value="THETAFLOW — CATALYST ANALYSIS").font = Font(name='Calibri', bold=True, size=14, color='1B2A4A')
    ws.cell(row=2, column=1, value=f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}").font = Font(name='Calibri', size=9, color='6B7D96')

    r = 4
    for lbl, val in [
        ("Headline", data.get("headline", "")),
        ("Direction", (top_picks[0].get("direction", "bullish") if top_picks else "").upper()),
        ("Conviction", top_picks[0].get("conviction", 0) if top_picks else 0),
        ("Time Horizon", data.get("ai_time_horizon", data.get("time_horizon", ""))),
    ]:
        write_label(ws, r, 1, lbl, bold=True)
        c = ws.cell(row=r, column=2, value=val)
        if lbl == "Conviction":
            c.number_format = pct_fmt
        r += 1

    r += 1
    write_label(ws, r, 1, "Investment Themes", bold=True)
    r += 1
    for ch in data.get("chains", []):
        ws.cell(row=r, column=1, value=ch.get("chain_name", ""))
        ws.cell(row=r, column=2, value=ch.get("relevance_score", 0))
        ws.cell(row=r, column=2).number_format = pct_fmt
        r += 1

    r += 1
    write_label(ws, r, 1, "Thesis", bold=True)
    r += 1
    ws.cell(row=r, column=1, value=data.get("summary", ""))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

    op = data.get("obvious_play")
    if op:
        r += 2
        write_label(ws, r, 1, "Obvious Play (Priced In)", bold=True)
        ws.cell(row=r, column=2, value=f"{op.get('ticker', '')} — {op.get('summary', '')}")

    r += 2
    write_label(ws, r, 1, "Risk Factors", bold=True)
    for rf in data.get("risk_factors", []):
        r += 1
        ws.cell(row=r, column=1, value=rf).font = label_font

    r += 2
    write_label(ws, r, 1, "Contrarian View", bold=True)
    r += 1
    ws.cell(row=r, column=1, value=data.get("contrarian_view", "")).font = label_font

    # ════════════════════════════════════════
    # SHEET 2: Revenue Buildup & P&L Pro Forma
    # ════════════════════════════════════════
    ws2 = wb.create_sheet("P&L Pro Forma")
    ws2.sheet_properties.tabColor = "34D399"
    ws2.column_dimensions['A'].width = 32

    # Each ticker gets a column block: Base Case (col) + Event Case (col+1)
    # Layout: Col A = labels, then 2 cols per ticker
    ws2.cell(row=1, column=1, value="INCOME STATEMENT — BASE vs EVENT-ADJUSTED").font = Font(name='Calibri', bold=True, size=12, color='1B2A4A')
    ws2.cell(row=2, column=1, value="Yellow cells are editable assumptions. Formulas auto-update.").font = Font(name='Calibri', italic=True, size=9, color='6B7D96')

    # Row labels for the P&L
    labels = [
        ("", False, False),  # row 4 = ticker headers
        ("Share Price", False, False),
        ("Shares Outstanding (M)", False, False),
        ("Market Cap ($M)", False, False),
        ("", False, False),
        ("REVENUE BUILDUP", True, False),  # row 9
        ("Revenue TTM ($M)", False, False),  # row 10 — INPUT
        ("Base Growth Rate", False, False),  # row 11 — INPUT
        ("Base Case Revenue ($M)", False, False),  # row 12 — FORMULA
        ("Event Revenue Impact ($M)", False, False),  # row 13 — INPUT
        ("Event-Adjusted Revenue ($M)", False, False),  # row 14 — FORMULA
        ("Event Revenue Growth", False, False),  # row 15 — FORMULA
        ("", False, False),
        ("PROFITABILITY", True, False),  # row 17
        ("Gross Margin", False, False),  # row 18 — INPUT
        ("Gross Profit ($M)", False, False),  # row 19 — FORMULA (base)
        ("EBITDA Margin", False, False),  # row 20 — INPUT
        ("EBITDA ($M)", False, False),  # row 21 — FORMULA
        ("D&A (% of Rev)", False, False),  # row 22 — INPUT
        ("EBIT ($M)", False, False),  # row 23 — FORMULA
        ("Tax Rate", False, False),  # row 24 — INPUT
        ("Net Income ($M)", False, False),  # row 25 — FORMULA
        ("Net Margin", False, False),  # row 26 — FORMULA
        ("", False, False),
        ("EARNINGS", True, False),  # row 28
        ("EPS (Base)", False, False),  # row 29 — FORMULA
        ("EPS (Event-Adjusted)", False, False),  # row 30 — FORMULA
        ("EPS Revision", False, False),  # row 31 — FORMULA
        ("", False, False),
        ("VALUATION", True, False),  # row 33
        ("Current P/E", False, False),  # row 34 — FORMULA
        ("Event P/E", False, False),  # row 35 — FORMULA
        ("Implied Price at Base P/E", False, False),  # row 36 — FORMULA
        ("Upside / Downside", False, False),  # row 37 — FORMULA
    ]

    START_ROW = 4
    for i, (lbl, bold, _) in enumerate(labels):
        if lbl:
            write_label(ws2, START_ROW + i, 1, lbl, bold=bold)

    # Write per-ticker columns
    for ti, t in enumerate(top_picks):
        fm = t.get("financial_model", {})
        bc = 2 + ti * 2  # base case column
        ec = bc + 1       # event case column
        bcl = CL(bc)
        ecl = CL(ec)
        ws2.column_dimensions[bcl].width = 18
        ws2.column_dimensions[ecl].width = 18

        r = START_ROW
        # Row 4: Headers
        ws2.cell(row=r, column=bc, value=f"{t.get('ticker', '')} Base").font = sec_font_sm
        ws2.cell(row=r, column=ec, value=f"{t.get('ticker', '')} Event").font = sec_font_sm
        ws2.cell(row=r, column=ec).font = Font(name='Calibri', bold=True, size=10, color='10B981' if t.get('direction') == 'bullish' else 'EF4444')

        price = t.get("current_price", 0) or 0
        shares = fm.get("shares_out_m", 0)
        rev = fm.get("revenue_ttm_m", 0)
        growth = (fm.get("revenue_growth_base_pct", 10) or 10) / 100
        ev_impact = fm.get("event_revenue_impact_m", 0)
        ev_margin_bps = fm.get("event_margin_impact_bps", 0)
        gm = (fm.get("gross_margin_pct", 50) or 50) / 100
        em = (fm.get("ebitda_margin_pct", 25) or 25) / 100
        da_pct = (fm.get("da_pct_rev", 5) or 5) / 100
        tax = (fm.get("tax_rate_pct", 21) or 21) / 100

        # Row 5: Share Price
        r += 1
        ws2.cell(row=r, column=bc, value=price).number_format = usd_fmt
        ws2.cell(row=r, column=ec, value=price).number_format = usd_fmt

        # Row 6: Shares Outstanding
        r += 1
        write_input(ws2, r, bc, shares, usd_m)
        ws2.cell(row=r, column=ec, value=f"={bcl}{r}").number_format = usd_m

        # Row 7: Market Cap = Price * Shares
        r += 1
        write_formula(ws2, r, bc, f"={bcl}{r-2}*{bcl}{r-1}", usd_m)
        write_formula(ws2, r, ec, f"={ecl}{r-2}*{ecl}{r-1}", usd_m)

        r += 1  # blank

        # Row 9: Revenue Buildup header
        r += 1

        # Row 10: Revenue TTM
        r += 1
        rev_row = r
        write_input(ws2, r, bc, rev, usd_m)
        ws2.cell(row=r, column=ec, value=f"={bcl}{r}").number_format = usd_m

        # Row 11: Base Growth Rate
        r += 1
        growth_row = r
        write_input(ws2, r, bc, growth, pct_fmt)
        ws2.cell(row=r, column=ec, value=f"={bcl}{r}").number_format = pct_fmt

        # Row 12: Base Case Revenue = TTM * (1 + growth)
        r += 1
        base_rev_row = r
        write_formula(ws2, r, bc, f"={bcl}{rev_row}*(1+{bcl}{growth_row})", usd_m, bold=True)
        # Event case: base + impact
        write_formula(ws2, r, ec, f"={bcl}{r}+{ecl}{r+1}", usd_m, bold=True)

        # Row 13: Event Revenue Impact
        r += 1
        ev_imp_row = r
        ws2.cell(row=r, column=bc, value=0).number_format = usd_m
        write_input(ws2, r, ec, ev_impact, usd_m)

        # Row 14: Event-Adjusted Revenue (duplicate of row 12 event col for clarity)
        r += 1
        ev_rev_row = r
        write_formula(ws2, r, bc, f"={bcl}{base_rev_row}", usd_m, bold=True)
        write_formula(ws2, r, ec, f"={ecl}{base_rev_row}", usd_m, bold=True)

        # Row 15: Event Revenue Growth
        r += 1
        write_formula(ws2, r, bc, f"={bcl}{growth_row}", pct_fmt)
        write_formula(ws2, r, ec, f"=({ecl}{ev_rev_row}/{bcl}{rev_row})-1", pct_fmt)

        r += 1  # blank

        # Row 17: Profitability header
        r += 1

        # Row 18: Gross Margin
        r += 1
        gm_row = r
        write_input(ws2, r, bc, gm, pct_fmt)
        # Event: base margin + margin impact (bps)
        write_input(ws2, r, ec, gm + ev_margin_bps / 10000, pct_fmt)

        # Row 19: Gross Profit = Revenue * Gross Margin
        r += 1
        write_formula(ws2, r, bc, f"={bcl}{ev_rev_row}*{bcl}{gm_row}", usd_m)
        write_formula(ws2, r, ec, f"={ecl}{ev_rev_row}*{ecl}{gm_row}", usd_m)

        # Row 20: EBITDA Margin
        r += 1
        em_row = r
        write_input(ws2, r, bc, em, pct_fmt)
        write_input(ws2, r, ec, em + ev_margin_bps / 20000, pct_fmt)  # half of margin impact flows to EBITDA

        # Row 21: EBITDA = Revenue * EBITDA Margin
        r += 1
        ebitda_row = r
        write_formula(ws2, r, bc, f"={bcl}{ev_rev_row}*{bcl}{em_row}", usd_m, bold=True)
        write_formula(ws2, r, ec, f"={ecl}{ev_rev_row}*{ecl}{em_row}", usd_m, bold=True)

        # Row 22: D&A % of Rev
        r += 1
        da_row = r
        write_input(ws2, r, bc, da_pct, pct_fmt)
        ws2.cell(row=r, column=ec, value=f"={bcl}{r}").number_format = pct_fmt

        # Row 23: EBIT = EBITDA - D&A
        r += 1
        ebit_row = r
        write_formula(ws2, r, bc, f"={bcl}{ebitda_row}-{bcl}{ev_rev_row}*{bcl}{da_row}", usd_m, bold=True)
        write_formula(ws2, r, ec, f"={ecl}{ebitda_row}-{ecl}{ev_rev_row}*{ecl}{da_row}", usd_m, bold=True)

        # Row 24: Tax Rate
        r += 1
        tax_row = r
        write_input(ws2, r, bc, tax, pct_fmt)
        ws2.cell(row=r, column=ec, value=f"={bcl}{r}").number_format = pct_fmt

        # Row 25: Net Income = EBIT * (1 - Tax)
        r += 1
        ni_row = r
        write_formula(ws2, r, bc, f"={bcl}{ebit_row}*(1-{bcl}{tax_row})", usd_m, bold=True)
        write_formula(ws2, r, ec, f"={ecl}{ebit_row}*(1-{ecl}{tax_row})", usd_m, bold=True)

        # Row 26: Net Margin = NI / Rev
        r += 1
        write_formula(ws2, r, bc, f"={bcl}{ni_row}/{bcl}{ev_rev_row}", pct_fmt)
        write_formula(ws2, r, ec, f"={ecl}{ni_row}/{ecl}{ev_rev_row}", pct_fmt)

        r += 1  # blank
        r += 1  # Earnings header

        # Row 29: EPS Base = NI / Shares
        r += 1
        eps_base_row = r
        shares_row = START_ROW + 2  # row 6
        write_formula(ws2, r, bc, f"={bcl}{ni_row}/{bcl}{shares_row}", usd_fmt, bold=True)
        write_formula(ws2, r, ec, f"={ecl}{ni_row}/{ecl}{shares_row}", usd_fmt, bold=True)

        # Row 30: EPS Event (same as ec col of row 29)
        r += 1
        write_formula(ws2, r, bc, f"={bcl}{eps_base_row}", usd_fmt)
        write_formula(ws2, r, ec, f"={ecl}{eps_base_row}", usd_fmt, bold=True)

        # Row 31: EPS Revision = Event EPS - Base EPS
        r += 1
        write_formula(ws2, r, bc, "", usd_fmt)
        write_formula(ws2, r, ec, f"={ecl}{eps_base_row}-{bcl}{eps_base_row}", usd_fmt, bold=True)

        r += 1  # blank
        r += 1  # Valuation header

        # Row 34: Current P/E = Price / Base EPS
        r += 1
        pe_row = r
        price_row = START_ROW + 1  # row 5
        write_formula(ws2, r, bc, f"=IF({bcl}{eps_base_row}>0,{bcl}{price_row}/{bcl}{eps_base_row},\"N/A\")", '0.0x')
        write_formula(ws2, r, ec, f"=IF({ecl}{eps_base_row}>0,{ecl}{price_row}/{ecl}{eps_base_row},\"N/A\")", '0.0x')

        # Row 35: Event P/E (forward)
        r += 1

        # Row 36: Implied Price = Event EPS * Base P/E
        r += 1
        write_formula(ws2, r, bc, "", usd_fmt)
        write_formula(ws2, r, ec, f"=IF(ISNUMBER({bcl}{pe_row}),{ecl}{eps_base_row}*{bcl}{pe_row},\"N/A\")", usd_fmt, bold=True)

        # Row 37: Upside/Downside = (Implied - Current) / Current
        r += 1
        implied_row = r - 1
        write_formula(ws2, r, bc, "", pct_fmt)
        write_formula(ws2, r, ec, f"=IF(ISNUMBER({ecl}{implied_row}),({ecl}{implied_row}-{ecl}{price_row})/{ecl}{price_row},\"N/A\")", pct_fmt, bold=True)

    auto_w(ws2)

    # ════════════════════════════════════════
    # SHEET 2b: Revenue Bridge (segment-level buildup)
    # ════════════════════════════════════════
    ws_rb = wb.create_sheet("Revenue Bridge")
    ws_rb.sheet_properties.tabColor = "06B6D4"
    ws_rb.column_dimensions['A'].width = 34

    ws_rb.cell(row=1, column=1, value="REVENUE BRIDGE — SEGMENT BUILDUP").font = Font(name='Calibri', bold=True, size=12, color='1B2A4A')
    ws_rb.cell(row=2, column=1, value="How the event impact flows through each business segment. Yellow cells are editable.").font = Font(name='Calibri', italic=True, size=9, color='6B7D96')

    rb_row = 4
    for ti, t in enumerate(top_picks):
        fm = t.get("financial_model", {})
        segments = fm.get("revenue_segments", [])
        ticker = t.get("ticker", "")

        # Ticker header
        ws_rb.cell(row=rb_row, column=1, value=f"{ticker} — {t.get('company', '')}").font = Font(name='Calibri', bold=True, size=11, color='1B2A4A')
        rb_row += 1

        # Column headers
        seg_headers = ["Segment", "TTM Revenue ($M)", "Base Growth", "Base Rev ($M)",
                       "Event Impact ($M)", "Impact Driver", "Event Rev ($M)", "Event Growth"]
        for c, h in enumerate(seg_headers, 1):
            ws_rb.cell(row=rb_row, column=c, value=h)
        hdr_row(ws_rb, rb_row, len(seg_headers))
        for c in range(2, 9):
            ws_rb.column_dimensions[CL(c)].width = 16
        ws_rb.column_dimensions[CL(6)].width = 40  # Impact Driver column wider
        header_r = rb_row

        if segments:
            for si, seg in enumerate(segments):
                sr = rb_row + 1 + si
                seg_name = seg.get("name", f"Segment {si+1}")
                seg_rev = seg.get("revenue_m", 0)
                seg_growth = (seg.get("growth_pct", 10) or 10) / 100
                seg_impact = seg.get("event_impact_m", 0)
                seg_driver = seg.get("impact_driver", "")

                ws_rb.cell(row=sr, column=1, value=seg_name).font = label_font
                write_input(ws_rb, sr, 2, seg_rev, usd_m)  # TTM Rev
                write_input(ws_rb, sr, 3, seg_growth, pct_fmt)  # Base Growth
                # Base Rev = TTM * (1+growth)
                write_formula(ws_rb, sr, 4, f"={CL(2)}{sr}*(1+{CL(3)}{sr})", usd_m)
                write_input(ws_rb, sr, 5, seg_impact, usd_m)  # Event Impact
                ws_rb.cell(row=sr, column=6, value=seg_driver).font = Font(name='Calibri', italic=True, size=9, color='4A5568')
                # Event Rev = Base + Impact
                write_formula(ws_rb, sr, 7, f"={CL(4)}{sr}+{CL(5)}{sr}", usd_m, bold=True)
                # Event Growth = Event Rev / TTM - 1
                write_formula(ws_rb, sr, 8, f"={CL(7)}{sr}/{CL(2)}{sr}-1", pct_fmt)
                for c in range(1, len(seg_headers) + 1):
                    ws_rb.cell(row=sr, column=c).border = thin_border

            # Totals row
            tr = rb_row + 1 + len(segments)
            ws_rb.cell(row=tr, column=1, value="TOTAL").font = sec_font_sm
            first_seg = rb_row + 1
            last_seg = rb_row + len(segments)
            for col_idx in [2, 4, 5, 7]:  # Sum columns: TTM Rev, Base Rev, Impact, Event Rev
                ws_rb.cell(row=tr, column=col_idx,
                           value=f"=SUM({CL(col_idx)}{first_seg}:{CL(col_idx)}{last_seg})")
                ws_rb.cell(row=tr, column=col_idx).number_format = usd_m
                ws_rb.cell(row=tr, column=col_idx).font = num_font_b
            # Total event growth = Total Event Rev / Total TTM - 1
            write_formula(ws_rb, tr, 8, f"={CL(7)}{tr}/{CL(2)}{tr}-1", pct_fmt, bold=True)
            for c in range(1, len(seg_headers) + 1):
                ws_rb.cell(row=tr, column=c).border = thick_border

            # Delta row: shows the change
            tr += 1
            ws_rb.cell(row=tr, column=1, value="DELTA (Event - Base)").font = Font(name='Calibri', bold=True, size=10, color='10B981')
            write_formula(ws_rb, tr, 4, f"={CL(7)}{tr-1}-{CL(4)}{tr-1}", usd_m, bold=True)
            ws_rb.cell(row=tr, column=4).font = green_font
            ws_rb.cell(row=tr, column=6, value=f"Sum of segment-level event impacts for {ticker}").font = Font(name='Calibri', italic=True, size=9, color='6B7D96')

            rb_row = tr + 3  # space before next ticker
        else:
            # No segments — show placeholder
            rb_row += 1
            ws_rb.cell(row=rb_row, column=1, value="No segment data available").font = Font(name='Calibri', italic=True, size=10, color='6B7D96')
            rev = fm.get("revenue_ttm_m", 0)
            ev_imp = fm.get("event_revenue_impact_m", 0)
            ws_rb.cell(row=rb_row, column=2, value=rev).number_format = usd_m
            ws_rb.cell(row=rb_row, column=5, value=ev_imp).number_format = usd_m
            rb_row += 3

    auto_w(ws_rb)

    # ════════════════════════════════════════
    # SHEET 3: DCF Valuation (top pick)
    # ════════════════════════════════════════
    if top_picks:
        t0 = top_picks[0]
        fm0 = t0.get("financial_model", {})
        ws3 = wb.create_sheet("DCF Model")
        ws3.sheet_properties.tabColor = "FBBF24"
        ws3.column_dimensions['A'].width = 30

        ws3.cell(row=1, column=1, value=f"DCF VALUATION — {t0.get('ticker', '')}").font = Font(name='Calibri', bold=True, size=12, color='1B2A4A')
        ws3.cell(row=2, column=1, value=f"{t0.get('company', '')}").font = Font(name='Calibri', size=10, color='4A5568')
        ws3.cell(row=3, column=1, value="Yellow cells are editable. All projections and valuation update automatically.").font = Font(name='Calibri', italic=True, size=9, color='6B7D96')

        # ── Assumptions block (rows 5-14) ──
        r = 5
        write_label(ws3, r, 1, "ASSUMPTIONS", bold=True)
        ws3.cell(row=r, column=2, value="Value").font = sec_font_sm

        rev0 = fm0.get("revenue_ttm_m", 10000) or 10000
        g0 = (fm0.get("revenue_growth_base_pct", 10) or 10) / 100
        em0 = (fm0.get("ebitda_margin_pct", 25) or 25) / 100
        capex0 = (fm0.get("capex_pct_rev", 8) or 8) / 100
        da0 = (fm0.get("da_pct_rev", 5) or 5) / 100
        tax0 = (fm0.get("tax_rate_pct", 21) or 21) / 100
        wacc0 = (fm0.get("wacc_pct", 10) or 10) / 100
        tg0 = (fm0.get("terminal_growth_pct", 3) or 3) / 100
        nd0 = fm0.get("net_debt_m", 0) or 0
        shares0 = fm0.get("shares_out_m", 100) or 100
        ev_impact0 = fm0.get("event_revenue_impact_m", 0) or 0
        price0 = t0.get("current_price", 0) or 0

        assumptions = [
            ("Revenue TTM ($M)", rev0, usd_m),
            ("Revenue Growth Yr 1 (incl. event)", g0, pct_fmt),
            ("Growth Decay / Yr", 0.02, pct_fmt),
            ("EBITDA Margin", em0, pct_fmt),
            ("Capex (% of Rev)", capex0, pct_fmt),
            ("D&A (% of Rev)", da0, pct_fmt),
            ("Tax Rate", tax0, pct_fmt),
            ("NWC (% of Rev Change)", 0.10, pct_fmt),
            ("WACC", wacc0, pct_fmt),
            ("Terminal Growth Rate", tg0, pct_fmt),
            ("Net Debt ($M)", nd0, usd_m),
            ("Shares Outstanding (M)", shares0, '#,##0'),
            ("Event Revenue Impact ($M)", ev_impact0, usd_m),
            ("Current Share Price", price0, usd_fmt),
        ]
        for i, (lbl, val, fmt) in enumerate(assumptions):
            arow = r + 1 + i
            write_label(ws3, arow, 1, lbl)
            write_input(ws3, arow, 2, val, fmt)

        # Named rows for formula references
        REV_ROW = r + 1       # B6
        G1_ROW = r + 2        # B7
        GDECAY_ROW = r + 3    # B8
        EM_ROW = r + 4        # B9
        CAPEX_ROW = r + 5     # B10
        DA_ROW = r + 6        # B11
        TAX_ROW = r + 7       # B12
        NWC_ROW = r + 8       # B13
        WACC_ROW = r + 9      # B14
        TG_ROW = r + 10       # B15
        ND_ROW = r + 11       # B16
        SH_ROW = r + 12       # B17
        EV_IMP_ROW = r + 13   # B18
        PRICE_ROW = r + 14    # B19

        # ── Projection table (rows 22+) ──
        pr = r + 1 + len(assumptions) + 2  # start of projection table
        write_label(ws3, pr, 1, "PROJECTION", bold=True)

        # Column headers: A=label, B=Year 0, C=Year 1, ... G=Year 5
        ws3.cell(row=pr, column=2, value="Year 0 (LTM)").font = sec_font_sm
        for y in range(1, PROJ_YEARS + 1):
            ws3.cell(row=pr, column=2 + y, value=f"Year {y}").font = sec_font_sm
            ws3.column_dimensions[CL(2 + y)].width = 16
        ws3.column_dimensions['B'].width = 16

        # Revenue row
        pr += 1
        rev_pr = pr
        write_label(ws3, pr, 1, "Revenue ($M)", bold=True)
        ws3.cell(row=pr, column=2, value=f"=$B${REV_ROW}").number_format = usd_m
        ws3.cell(row=pr, column=2).font = num_font_b
        for y in range(1, PROJ_YEARS + 1):
            col = 2 + y
            pcol = CL(col - 1)
            # Year 1 growth = G1, subsequent years decay by GDECAY
            if y == 1:
                ws3.cell(row=pr, column=col, value=f"={pcol}{pr}*(1+$B${G1_ROW})").number_format = usd_m
            else:
                ws3.cell(row=pr, column=col, value=f"={pcol}{pr}*(1+MAX($B${G1_ROW}-$B${GDECAY_ROW}*{y-1},0.02))").number_format = usd_m
            ws3.cell(row=pr, column=col).font = num_font_b

        # Growth rate row
        pr += 1
        write_label(ws3, pr, 1, "Revenue Growth", indent=1)
        ws3.cell(row=pr, column=2, value="")
        for y in range(1, PROJ_YEARS + 1):
            col = 2 + y
            pcol = CL(col - 1)
            ws3.cell(row=pr, column=col, value=f"={CL(col)}{rev_pr}/{pcol}{rev_pr}-1").number_format = pct_fmt

        # EBITDA
        pr += 1
        ebitda_pr = pr
        write_label(ws3, pr, 1, "EBITDA ($M)", bold=True)
        for y in range(0, PROJ_YEARS + 1):
            col = 2 + y
            ws3.cell(row=pr, column=col, value=f"={CL(col)}{rev_pr}*$B${EM_ROW}").number_format = usd_m
            ws3.cell(row=pr, column=col).font = num_font_b

        # EBITDA Margin (reference row)
        pr += 1
        write_label(ws3, pr, 1, "EBITDA Margin", indent=1)
        for y in range(0, PROJ_YEARS + 1):
            col = 2 + y
            ws3.cell(row=pr, column=col, value=f"=$B${EM_ROW}").number_format = pct_fmt

        # D&A
        pr += 1
        da_pr = pr
        write_label(ws3, pr, 1, "(-) D&A ($M)", indent=1)
        for y in range(0, PROJ_YEARS + 1):
            col = 2 + y
            ws3.cell(row=pr, column=col, value=f"={CL(col)}{rev_pr}*$B${DA_ROW}").number_format = usd_m

        # EBIT
        pr += 1
        ebit_pr = pr
        write_label(ws3, pr, 1, "EBIT ($M)", bold=True)
        for y in range(0, PROJ_YEARS + 1):
            col = 2 + y
            ws3.cell(row=pr, column=col, value=f"={CL(col)}{ebitda_pr}-{CL(col)}{da_pr}").number_format = usd_m
            ws3.cell(row=pr, column=col).font = num_font_b

        # Taxes
        pr += 1
        tax_pr = pr
        write_label(ws3, pr, 1, "(-) Taxes ($M)", indent=1)
        for y in range(0, PROJ_YEARS + 1):
            col = 2 + y
            ws3.cell(row=pr, column=col, value=f"={CL(col)}{ebit_pr}*$B${TAX_ROW}").number_format = usd_m

        # NOPAT
        pr += 1
        nopat_pr = pr
        write_label(ws3, pr, 1, "NOPAT ($M)")
        for y in range(0, PROJ_YEARS + 1):
            col = 2 + y
            ws3.cell(row=pr, column=col, value=f"={CL(col)}{ebit_pr}-{CL(col)}{tax_pr}").number_format = usd_m

        # Add back D&A
        pr += 1
        addback_pr = pr
        write_label(ws3, pr, 1, "(+) D&A ($M)", indent=1)
        for y in range(0, PROJ_YEARS + 1):
            col = 2 + y
            ws3.cell(row=pr, column=col, value=f"={CL(col)}{da_pr}").number_format = usd_m

        # Capex
        pr += 1
        capex_pr = pr
        write_label(ws3, pr, 1, "(-) Capex ($M)", indent=1)
        for y in range(0, PROJ_YEARS + 1):
            col = 2 + y
            ws3.cell(row=pr, column=col, value=f"={CL(col)}{rev_pr}*$B${CAPEX_ROW}").number_format = usd_m

        # NWC Change
        pr += 1
        nwc_pr = pr
        write_label(ws3, pr, 1, "(-) NWC Change ($M)", indent=1)
        ws3.cell(row=pr, column=2, value=0).number_format = usd_m
        for y in range(1, PROJ_YEARS + 1):
            col = 2 + y
            pcol = CL(col - 1)
            ws3.cell(row=pr, column=col, value=f"=({CL(col)}{rev_pr}-{pcol}{rev_pr})*$B${NWC_ROW}").number_format = usd_m

        # ── UFCF (Unlevered Free Cash Flow) ──
        pr += 1
        ufcf_pr = pr
        write_label(ws3, pr, 1, "UNLEVERED FCF ($M)", bold=True)
        for y in range(0, PROJ_YEARS + 1):
            col = 2 + y
            ws3.cell(row=pr, column=col,
                     value=f"={CL(col)}{nopat_pr}+{CL(col)}{addback_pr}-{CL(col)}{capex_pr}-{CL(col)}{nwc_pr}").number_format = usd_m
            ws3.cell(row=pr, column=col).font = num_font_b
            ws3.cell(row=pr, column=col).border = thick_border

        # ── Discount factors ──
        pr += 2
        disc_pr = pr
        write_label(ws3, pr, 1, "Discount Factor")
        for y in range(1, PROJ_YEARS + 1):
            col = 2 + y
            ws3.cell(row=pr, column=col, value=f"=1/(1+$B${WACC_ROW})^{y}").number_format = '0.0000'

        # PV of FCF
        pr += 1
        pv_pr = pr
        write_label(ws3, pr, 1, "PV of FCF ($M)")
        for y in range(1, PROJ_YEARS + 1):
            col = 2 + y
            ws3.cell(row=pr, column=col, value=f"={CL(col)}{ufcf_pr}*{CL(col)}{disc_pr}").number_format = usd_m

        # ── Valuation summary (below projection table) ──
        pr += 2
        vs = pr
        write_label(ws3, vs, 1, "VALUATION SUMMARY", bold=True)

        # Sum of PV of FCFs
        vs += 1
        write_label(ws3, vs, 1, "Sum of PV of FCFs ($M)")
        last_col = CL(2 + PROJ_YEARS)
        write_formula(ws3, vs, 2, f"=SUM(C{pv_pr}:{last_col}{pv_pr})", usd_m, bold=True)
        sum_pv_row = vs

        # Terminal Value = Final Year FCF * (1+g) / (WACC - g)
        vs += 1
        write_label(ws3, vs, 1, "Terminal Value ($M)")
        write_formula(ws3, vs, 2, f"={last_col}{ufcf_pr}*(1+$B${TG_ROW})/($B${WACC_ROW}-$B${TG_ROW})", usd_m, bold=True)
        tv_row = vs

        # PV of Terminal Value
        vs += 1
        write_label(ws3, vs, 1, "PV of Terminal Value ($M)")
        write_formula(ws3, vs, 2, f"=B{tv_row}/(1+$B${WACC_ROW})^{PROJ_YEARS}", usd_m, bold=True)
        pv_tv_row = vs

        # Enterprise Value
        vs += 1
        write_label(ws3, vs, 1, "Enterprise Value ($M)", bold=True)
        write_formula(ws3, vs, 2, f"=B{sum_pv_row}+B{pv_tv_row}", usd_m, bold=True)
        ws3.cell(row=vs, column=2).border = thick_border
        ev_row = vs

        # Less Net Debt
        vs += 1
        write_label(ws3, vs, 1, "(-) Net Debt ($M)")
        write_formula(ws3, vs, 2, f"=$B${ND_ROW}", usd_m)

        # Equity Value
        vs += 1
        write_label(ws3, vs, 1, "Equity Value ($M)", bold=True)
        write_formula(ws3, vs, 2, f"=B{ev_row}-B{vs-1}", usd_m, bold=True)
        eq_row = vs

        # Implied Share Price
        vs += 1
        write_label(ws3, vs, 1, "Implied Share Price", bold=True)
        write_formula(ws3, vs, 2, f"=B{eq_row}/$B${SH_ROW}", usd_fmt, bold=True)
        ws3.cell(row=vs, column=2).font = Font(name='Consolas', bold=True, size=12, color='1B2A4A')
        ws3.cell(row=vs, column=2).border = thick_border
        impl_row = vs

        # Current Price
        vs += 1
        write_label(ws3, vs, 1, "Current Price")
        write_formula(ws3, vs, 2, f"=$B${PRICE_ROW}", usd_fmt)

        # Upside/Downside
        vs += 1
        write_label(ws3, vs, 1, "UPSIDE / DOWNSIDE", bold=True)
        write_formula(ws3, vs, 2, f"=(B{impl_row}-B{vs-1})/B{vs-1}", pct_fmt, bold=True)
        ws3.cell(row=vs, column=2).font = Font(name='Consolas', bold=True, size=12, color='10B981')

        auto_w(ws3)

    # ════════════════════════════════════════
    # SHEET 4: Trades Overview
    # ════════════════════════════════════════
    ws4 = wb.create_sheet("Trades")
    ws4.sheet_properties.tabColor = "818CF8"
    headers = ["Ticker", "Company", "Action", "Direction", "Conviction",
               "Order", "Chain", "Layer", "Impact Score", "Risk/Reward",
               "Earnings Impact", "Precedent", "Thesis"]
    for c, h in enumerate(headers, 1):
        ws4.cell(row=1, column=c, value=h)
    hdr_row(ws4, 1, len(headers))

    for i, t in enumerate(all_picks, 2):
        ws4.cell(row=i, column=1, value=t.get("ticker", "")).font = num_font_b
        ws4.cell(row=i, column=2, value=t.get("company", ""))
        ws4.cell(row=i, column=3, value=t.get("action", ""))
        ws4.cell(row=i, column=4, value=t.get("direction", ""))
        ws4.cell(row=i, column=5, value=t.get("conviction", 0)).number_format = pct_fmt
        ws4.cell(row=i, column=6, value=t.get("order", 1))
        ws4.cell(row=i, column=7, value=t.get("chain_name", ""))
        ws4.cell(row=i, column=8, value=t.get("layer", ""))
        ws4.cell(row=i, column=9, value=t.get("impact_score", 0))
        ws4.cell(row=i, column=10, value=t.get("risk_reward", ""))
        ws4.cell(row=i, column=11, value=t.get("earnings_impact", ""))
        ws4.cell(row=i, column=12, value=t.get("precedent_move", ""))
        ws4.cell(row=i, column=13, value=t.get("thesis", t.get("investment_analysis", "")))
        for c in range(1, len(headers) + 1):
            ws4.cell(row=i, column=c).border = thin_border

    auto_w(ws4)

    # ════════════════════════════════════════
    # SHEET 5: Portfolio Sizing (if provided)
    # ════════════════════════════════════════
    portfolio_size = data.get("portfolio_size")
    if portfolio_size and top_picks:
        ws5 = wb.create_sheet("Portfolio Sizing")
        ws5.sheet_properties.tabColor = "34D399"
        ws5.cell(row=1, column=1, value="PORTFOLIO SIZING").font = Font(name='Calibri', bold=True, size=12, color='1B2A4A')
        write_label(ws5, 2, 1, "Portfolio Value")
        write_input(ws5, 2, 2, portfolio_size, usd_whole)

        ps_h = ["Ticker", "Company", "Action", "Conviction", "Weight",
                "Dollar Allocation", "Shares", "Price", "Proj Return", "Dollar P&L"]
        for c, h in enumerate(ps_h, 1):
            ws5.cell(row=4, column=c, value=h)
        hdr_row(ws5, 4, len(ps_h))

        valid = [t for t in top_picks if t.get("current_price", 0) > 0]
        tc = sum(t.get("conviction", 0.5) for t in valid) or 1
        for i, t in enumerate(valid, 5):
            cv = t.get("conviction", 0.5)
            w = cv / tc
            d = portfolio_size * w
            pr = t.get("current_price", 0)
            sh = int(d / pr) if pr else 0
            ps = 1 if t.get("direction") == "bullish" else -1
            ret = math.exp(ps * cv * 0.003 * 22) - 1
            pnl = d * ret
            ws5.cell(row=i, column=1, value=t.get("ticker", "")).font = num_font_b
            ws5.cell(row=i, column=2, value=t.get("company", ""))
            ws5.cell(row=i, column=3, value=t.get("action", ""))
            ws5.cell(row=i, column=4, value=cv).number_format = pct_fmt
            ws5.cell(row=i, column=5, value=w).number_format = pct_fmt
            ws5.cell(row=i, column=6, value=d).number_format = usd_whole
            ws5.cell(row=i, column=7, value=sh)
            ws5.cell(row=i, column=8, value=pr).number_format = usd_fmt
            ws5.cell(row=i, column=9, value=ret).number_format = pct_fmt
            ws5.cell(row=i, column=10, value=pnl).number_format = usd_whole
        auto_w(ws5)

    # ── Write to bytes and return ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    headline_slug = re.sub(r'[^a-zA-Z0-9]', '_', data.get("headline", "analysis")[:30]).strip('_')
    filename = f"thetaflow_{headline_slug}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )


# ── Auth Endpoints ──

@app.route("/api/auth/register", methods=["POST"])
def register():
    data = request.json or {}
    email = data.get("email", "").strip().lower()
    password = data.get("password", "")
    if not email or not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
        return jsonify({"error": "Valid email required"}), 400
    if len(password) < 6:
        return jsonify({"error": "Password must be at least 6 characters"}), 400
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.execute("INSERT INTO users (email, password_hash, tier) VALUES (?, ?, ?)",
                     (email, hash_password(password), "free"))
        conn.commit()
        user_id = conn.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone()[0]
        conn.close()
        session['user_id'] = user_id
        return jsonify({"success": True, "user": {"id": user_id, "email": email, "tier": "free"}})
    except sqlite3.IntegrityError:
        return jsonify({"error": "Email already registered"}), 409

@app.route("/api/auth/login", methods=["POST"])
def login():
    data = request.json or {}
    email = data.get("email", "").strip().lower()
    password = data.get("password", "")
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    row = conn.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
    if not row or not verify_password(password, row["password_hash"]):
        conn.close()
        return jsonify({"error": "Invalid email or password"}), 401
    conn.execute("UPDATE users SET last_login_at = ? WHERE id = ?",
                 (datetime.utcnow().isoformat(), row["id"]))
    conn.commit()
    conn.close()
    user = dict(row)
    session['user_id'] = user["id"]
    return jsonify({"success": True, "user": {
        "id": user["id"], "email": user["email"], "tier": user["tier"],
        "api_key": user["api_key"],
    }})

@app.route("/api/auth/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"success": True})

@app.route("/api/auth/me", methods=["GET"])
def get_me():
    if 'user_id' not in session:
        return jsonify({"authenticated": False}), 401
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    row = conn.execute("SELECT * FROM users WHERE id = ?", (session['user_id'],)).fetchone()
    conn.close()
    if not row:
        session.clear()
        return jsonify({"authenticated": False}), 401
    user = dict(row)
    return jsonify({"authenticated": True, "user": {
        "id": user["id"], "email": user["email"], "tier": user["tier"],
        "api_key": user["api_key"],
    }})

# ── Startup & Scheduler ──

def initialize_data():
    global last_collection_at
    conn = sqlite3.connect(DB_PATH)
    count = conn.execute("SELECT COUNT(*) FROM events").fetchone()[0]
    conn.close()
    if count == 0:
        logger.info("Empty database — running initial event collection...")
        orchestrator.run_collection()
        last_collection_at = datetime.utcnow().isoformat()
        logger.info("Initial data loaded.")

def start_scheduler():
    global last_collection_at
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        sched = BackgroundScheduler()

        def scheduled_collection():
            global last_collection_at
            try:
                logger.info("Scheduled event collection starting...")
                orchestrator.run_collection()
                last_collection_at = datetime.utcnow().isoformat()
                logger.info(f"Scheduled collection complete at {last_collection_at}")
            except Exception as e:
                logger.error(f"Scheduled collection failed: {e}")

        sched.add_job(scheduled_collection, 'interval', hours=4,
                      id='event_collection', replace_existing=True)
        sched.start()
        logger.info("Scheduler started: event collection every 4 hours")
    except ImportError:
        logger.warning("APScheduler not installed — scheduled jobs disabled")

if __name__ == "__main__":
    initialize_data()
    if os.environ.get("WERKZEUG_RUN_MAIN") or not app.debug:
        start_scheduler()
    port = int(os.getenv("PORT", 5002))
    app.run(host="0.0.0.0", port=port, debug=True)
